import math
from playwright.sync_api import Playwright, sync_playwright
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import os
import pandas as pd
from datetime import datetime
import config as cf #configuration file will be generated once the application form is filled

#Variables
fname = cf.fname
lname = cf.lname
phno = cf.phno
EMAIL = cf.email
PASSWORD = cf.pw
url = "https://www.linkedin.com/login"
role = cf.role
location = cf.loc
yoe = cf.yoe
#resume path
rname = os.getcwd()+"/"+cf.resume
head = cf.head
all_job_ids = []
filename = "LinkedInJobs.xlsx"
city = cf.city
status_dict = {}  # Global dictionary

def update_status_dict(job_url, status):
    """Updates the global status dictionary with the given job URL and status."""
    global status_dict
    status_dict[job_url.strip()] = status  # Strip spaces for consistency
    print(f"Updating {job_url} to '{status}'in dictionary")

def single_form(page):
    if page.locator('input[name="firstName"]').is_visible():
        page.locator('input[name="firstName"]').fill(fname)
        print("First Name updated")

    if page.locator('input[name="lastName"]').is_visible():
        page.locator('input[name="lastName"]').fill(lname)
        print("Last Name updated")

    if page.locator('input[name="phoneNumber"]').is_visible():
        page.locator('input[name="phoneNumber"]').fill(phno)
        print("Phone number updated")

    # if page.get_by_label("Mobile phone number").is_visible():
    #     page.get_by_label("Mobile phone number").fill(phno)
    #     print(f"Mobile Number filled with {phno}")
    if page.get_by_label("Location (city)").is_visible():
        page.get_by_label("Location (city)").fill(city)
        print(f"City field filled with {city}")

    if page.locator("label").filter(has_text="Upload resume").is_visible():
        with page.expect_file_chooser() as fc_info:
            page.locator("label").filter(has_text="Upload resume").click()
        file_chooser = fc_info.value
        file_chooser.set_files(rname)
        print("Resume uploaded successfully")

def update_excel_from_dict(filename):
    """Updates the Excel file with all statuses from the global status dictionary."""
    global status_dict
    try:
        df = pd.read_excel(filename)

        # Standardize column names
        df.columns = df.columns.str.strip()
        df['Job URL'] = df['Job URL'].astype(str).str.strip()  # Ensure consistency

        # Ensure 'Applied Status' column exists
        if 'Applied Status' not in df.columns:
            df['Applied Status'] = pd.NA

        df['Applied Status'] = df['Applied Status'].astype(str)

        # Debugging: Print job URLs before updating
        print("Existing job URLs in DataFrame:", df['Job URL'].tolist())

        # Update statuses in DataFrame
        for job_url, status in status_dict.items():
            if job_url in df['Job URL'].values:
                print(f"Updating {job_url} to '{status}' in excel sheet.")
                df.loc[df['Job URL'] == job_url, 'Applied Status'] = status
            else:
                print(f"Job URL {job_url} not found in DataFrame!")

        # Save the DataFrame back to Excel
        df.to_excel(filename, index=False)
        print("Excel file updated successfully with all statuses.")
    except Exception as e:
        print(f"Error updating Excel: {e}")

def fill_form_until_review(page):
    review_attempts = 0 # Counter for repeated review page encounters
    fill_form_attempts = 0
    job_url = page.url  # Capture the current job URL
    while True:
        print("in fill_form_until_review function")
        try:
            fill_form(page)  # Fill the form fields
        except Exception as e:
            print(f"Exception: {e}")
            update_status_dict(job_url,"Attempted, but failed")
            break

        try:
            # Click "Continue to next step" if not on the review page
            next_step_button = page.get_by_label("Continue to next step")
            print("Next button present")
            if next_step_button.is_visible():
                next_step_button.click()
                print("Clicked 'Continue to next step'. Moving forward.")
                page.wait_for_timeout(2000)  # Small wait to let page load
            else:
                # Check if "Review your application" button is present
                review_button = page.get_by_label("Review your application")
                print("Review button present")
                if review_button.is_visible():
                    print("Reached the review page. Stopping form filling.")
                    print("Clicking review button.")
                    review_button.click()
                    page.wait_for_timeout(3000)
                    review_attempts += 1  # Increment review page counter
                    if review_attempts >= 3:  # If stuck on the review page
                        print("Stuck on review page. Exiting loop and updating status.")
                        update_status_dict(job_url,"Attempted, but failed")
                        break
                    # Check for submit button
                    submit_button = page.get_by_label("Submit Application")
                    if submit_button.is_visible():
                        print("Submit Button present. Clicking submit button!")
                        submit_button.click()
                        update_status_dict(job_url,"Submitted")  # Update the status to "Submitted"
                        break  # Exit the loop after submission
                else:
                    print("No action button present")
                    update_status_dict(job_url,"Attempted, but failed")
                    break  # Exit the loop if no next step or review button exists

        except Exception as e:
            print(f"Error encountered: {e}. Ending process.")
            break  # Stop if any error occurs (e.g., button not found)

def fill_form(page):
    job = page.url
    labels = page.locator("label").all()  # Get all labels

    for label in labels:
        input_id = label.get_attribute("for")  # Get the associated input field ID

        if input_id:
            input_element = page.locator(f"xpath=//*[@id='{input_id}']")
            input_type = input_element.get_attribute("type")
            tag_name = input_element.evaluate("(el) => el.tagName.toLowerCase()")  # Get tag name

            if input_element.count() == 0:
                print(f"No input field found for label '{label.text_content().strip()}'. Skipping.")
                continue  # Skip if no input found

            try:
                if input_type == "text":
                    input_element.fill(str(yoe))  # Fill text fields with "yoe"
                    print(f"Filled '{label.inner_text().strip()}' with {yoe}")

                elif input_type == "radio":
                    if label.inner_text().strip().lower() == "yes":
                        try:
                            label.click()  # Click on the label
                            print(f"Clicked on label: {label.inner_text().strip()}")
                        except Exception as e:
                            print(f"Could not click on label '{label.inner_text().strip()}': {e}")

                elif tag_name == "select":
                    try:
                        input_element.select_option("Yes",timeout=1000)  # Try selecting "Yes"
                        print(f"Selected 'Yes' for '{label.inner_text().strip()}'")
                    except Exception:
                        #print(f"Skipped '{label.inner_text().strip()}' (Option 'Yes' not found)")
                        #update_status_in_excel(job, filename, "Attempted, but failed")
                        print("Handling Model Interferences ...")
                        return

                elif input_type == "checkbox":
                    # Skip if an overlay is intercepting pointer events
                    if input_element.is_visible() and input_element.is_enabled():
                        try:
                            # Check if an overlay is blocking the click
                            overlay = page.locator("#ember211")  # Modify the selector if needed

                            if overlay.is_visible():
                                print(f"Skipping checkbox '{label.inner_text().strip()}' due to modal interferences.")
                            else:
                                # Ensure it's not already checked before clicking
                                if input_element.get_attribute("aria-checked") != "true":
                                    input_element.scroll_into_view_if_needed()  # Ensure it's in view
                                    input_element.check(timeout=1000)
                                    print(f"Checked '{label.inner_text().strip()}'")
                                else:
                                    print(f"Checkbox '{label.inner_text().strip()}' is already checked, skipping.")
                        except Exception as e:
                            #print(f"Could not check '{label.inner_text().strip()}': {e}")
                            #update_status_in_excel(job, filename, "Attempted, but failed")
                            print("Checking Modal Interferences and skipping them.")
                else:
                    print("Oops! Something went wrong.")
                    update_status_dict(job,"Attempted, but failed")
                    return #break and go to the next job
            except Exception as e:
                print(f"Couldn't process: '{label.inner_text().strip()}', due to error- {e}")
                update_status_dict(job,"Attempted, but failed")
        else:
            print(f"Label '{label.inner_text().strip()}' has no 'for' attribute")
            update_status_dict(job,"Attempted, but failed")
            return  # Instead of breaking, continue processing other labels

def applyJobs(page):
    print("In Applying for jobs by reading from excel")

    # Reading job URL from excel and storing in an array
    workbook = load_workbook(filename)
    worksheet = workbook.active
    job_id_url = []

    # Collecting URLs for jobs where HR ID is None
    for i, row in enumerate(worksheet):
        if i == 0:  # Skips Column Heading
            continue
        url = row[1].value
        hr_id = row[3].value
        applied_status = row[4].value

        # Handle None or empty applied status
        if applied_status is None:
            applied_status = ""
        else:
            applied_status = str(applied_status).strip().lower()

        # Skip jobs with status 'Submitted' or 'Attempted, but failed'
        if hr_id is not None:
            print(f"  Skipping {url} - Status: {applied_status}")
            continue
        job_id_url.append(url)

    print("Job URL to apply:")
    k = 1

    for j in job_id_url:
        print({k}, " :- ", {j})
        k += 1

    print("---------Application Process Started---------")
    i = 0

    # Iterate over each job URL
    for job in job_id_url:
        error_occurred = False  # Initialize the flag at the start of each job processing
        i += 1
        print(f"----------{i}----------")
        print(f"In page url:{job}")

        # Opening a page and applying
        page.goto(job)
        page.wait_for_timeout(5000)

        recruiter = page.get_by_role("heading", name="Meet the hiring team").is_visible()
        if recruiter:
            print("Recruiter is present")
            recruiter_link = page.locator(".job-details-people-who-can-help__section a[data-test-app-aware-link]").first.get_attribute("href")
        else:
            print("No recruiter")
            recruiter_link = 0

        print(recruiter_link)
        # Update the HR ID in the Excel file for the matching job URL
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):  # Starting from row 2 to skip headers
            if row[1].value == job:  # Match Job URL with current page URL
                row[3].value = recruiter_link  # Update HR ID (4th column)
        workbook.save("LinkedInJobs.xlsx")  # Save changes immediately
        print(f"Updated Recruiter Information in the file: {filename}")

        # Clicking on the Easy Apply button
        try:
            easyApply = page.locator('button[aria-label^="Easy Apply to"]').nth(1)
            easyApply.scroll_into_view_if_needed()
            try:
                easyApply.click()
            except Exception as e:
                print(f"Error processing {job} due to: {e}")
                update_status_dict(job,"Attempted, but failed")
                continue  # Moves to the next job

            page.wait_for_timeout(5000)

            dialog_box = page.get_by_role("heading", name="Contact info").is_visible()
            if dialog_box:
                print("opened the dialog box")
                submit_button = page.get_by_label("Submit Application")
                next_step_button = page.get_by_label("Continue to next step")
                if submit_button.is_visible():
                    print("One Pager Form")
                    try:
                        single_form(page)
                        submit_button.click()
                        update_status_dict(job,"Submitted")  # Update the status to "Submitted"
                        continue  # Exit the loop after submission
                    except Exception as e:
                        print(f"Couldn't submit single page form due to {e}")
                        error_occurred = True

                elif next_step_button.is_visible():
                    print("Long Form")
                    if page.locator('input[name="firstName"]').is_visible():
                        page.locator('input[name="firstName"]').fill(fname)
                        print("First Name updated")

                    if page.locator('input[name="lastName"]').is_visible():
                        page.locator('input[name="lastName"]').fill(lname)
                        print("Last Name updated")

                    if page.get_by_label("Mobile phone number").is_visible():
                        page.get_by_label("Mobile phone number").fill(phno)
                        print(f"Mobile Number filled with {phno}")

                    if page.get_by_label("Location (city)").is_visible():
                        page.get_by_label("Location (city)").fill(city)
                        print(f"City field filled with {city}")

                    page.get_by_label("Continue to next step").click()

                    with page.expect_file_chooser() as fc_info:
                        page.locator("label").filter(has_text="Upload resume").click()  # Click upload button
                    file_chooser = fc_info.value  # Capture the file chooser event
                    file_chooser.set_files(rname)  # Set the file for upload
                    print("Resume uploaded successfully")
                    page.wait_for_timeout(5000)
                    review_button = page.get_by_label("Review your application")
                    next_step_button = page.get_by_label("Continue to next step")
                    if review_button.is_visible():
                        print("Review button is present after uploading resume")
                        print("Reached the review page. Stopping form filling.")
                        print("Clicking review button.")
                        review_button.click()
                        page.wait_for_timeout(3000)
                        review_attempts += 1  # Increment review page counter
                        if review_attempts >= 3:  # If stuck on the review page
                            print("Stuck on review page. Exiting loop and updating status.")
                            update_status_dict(job,"Attempted, but failed")
                        break
                        # Check for submit button
                        submit_button = page.get_by_label("Submit Application")
                        if submit_button.is_visible():
                            print("Submit Button present. Clicking submit button!")
                            submit_button.click()
                            update_status_dict(job,"Submitted")  # Update the status to "Submitted"
                            break  # Exit the loop after submission

                    if next_step_button.is_visible():
                        page.get_by_label("Continue to next step").click()
                        print("Next button is present after uploading resume.")
                        page.wait_for_timeout(5000)
                        try:
                            fill_form_until_review(page)
                        except Exception as e:
                            print(f"Couldn't fill form until the end due to {e}")
                            error_occurred = True
            else:
                print("Error in opening the dialog box after clicking on Easy Apply Button.")
                error_occurred = True

            page.wait_for_timeout(5000)

        except Exception as e:
            print("Error: Easy Apply Button not found.")
            error_occurred = True

        # Update the status only once after the application process for each job
        if error_occurred:
            update_status_dict(job,"Attempted, but failed")

    # Final Save and Close the Workbook
    workbook.save("LinkedInJobs.xlsx")# Save changes at the end of all processing
    workbook.close()  # Explicitly close the workbook
    print("Final save after applying for all jobs.")

def dupe_remove(filename):
   # Getting filepath of the workbook
   cwd = os.getcwd()
   filepath = os.path.join(cwd, filename)
   print("File path:", filepath)

   # Create DataFrame from the Excel file
   df = pd.read_excel(filepath)

   # Remove duplicates based on 'Job ID' and 'Job URL' columns
   df.drop_duplicates(subset=['Job ID', 'Job URL'], inplace=True)

   # Reset index after removing duplicates
   df.reset_index(drop=True, inplace=True)

   # Print the updated DataFrame and its count
   print("Updated DataFrame after removing duplicates:")
   print(df.to_string())

   # Write the updated DataFrame back to the Excel file
   df.to_excel(filepath, index=False)
   print("Updated data has been saved to the Excel file.")

def write_to_excel():
   """
    Writes extracted job IDs and related data into an Excel sheet.
   """
   # print(all_job_ids)
   # print(len(all_job_ids))
   sheet_name = "JobData"

   #Check if file already exists
   if os.path.exists(filename):
      workbook = load_workbook(filename)
      sheet = workbook.active  # Access the default sheet
      print("File exists already")
   else:
      #Create a new workboook
      workbook = Workbook()
      # Rename the default first sheet to 'JobData'
      sheet = workbook.active  # Access the default sheet
      sheet.title = "JobData"
      #Columns to be added
      sheet.append(["Job ID","Job URL", "Date","HR ID", "Applied Status"])

   # # Check if the sheet exists, otherwise create it
   # if sheet_name in workbook.sheetnames:
   #    sheet = workbook[sheet_name]

   today_date = datetime.today().strftime('%Y-%m-%d')

   #Writing into the sheet
   for job_id in all_job_ids:
      joburl = f"https://www.linkedin.com/jobs/view/{job_id}/"
      sheet.append([job_id,joburl,today_date,None,None])

   #save the workbook
   workbook.save(filename)
   #print(f"Saved {filename}")
   dupe_remove(filename)

def extract_job_ids(page):
   #Code for extracting job ids-
   paginated_url = page.url
   print(f"Paginated URL : {paginated_url}")
   page.goto(paginated_url)
   #Page load time
   page.wait_for_timeout(10000)
   # Wait for the job list container to load
   page.wait_for_selector('div.scaffold-layout__list', timeout=10000)
   # Get job ids of jobs listed in the current page
   try:
      html_response = page.content()
      soup = BeautifulSoup(html_response, 'html.parser')

      # Select job items using a stable structure
      job_items = soup.find_all('li', attrs={'data-occludable-job-id': True})

      if not job_items:
         print("No jobs found on the current page.")

      # Extract and print job IDs
      for job in job_items:
         job_ids = job.get('data-occludable-job-id')
         #print(job_ids)
         all_job_ids.append(job_ids)

   except Exception as e:
      print(f"Error while fetching job_ids: {e}")
      return

   #Print the total number of job ids extracted for the given user inputs
   # print(f"Total extracted job IDs: {len(all_job_ids)}")
   # for job_id in all_job_ids:
   #    print(job_id)

def extract_jobs(page) -> None:
   """
   Extract job links from the current page and print them.
   :param page: Playwright Page object
   """
   #print("Extracting jobs into an Excel sheet")

   #target_url = page.url+"&start={}"
   #print(f"Target URL: {target_url}")

   # Get the number of jobs from the page
   try:
      text = page.locator("small span").inner_text()
      number_of_jobs_filtered = int(text.split(" ")[0].replace(",", "")) # e.g : 1,006 -> 1006
      print(f"Number of jobs filtered: {number_of_jobs_filtered}")
   except Exception as e:
      print(f"Error while fetching the number of jobs: {e}")
      return


   number_of_pages = math.ceil(number_of_jobs_filtered/25)
   print(f"Number of pages : {number_of_pages}")

   #Code for hitting each page url:
   print("Hitting each url attempt")

   #wait for buttons to load
   #page.wait_for_selector("li[data-test-pagination-page-btn]", timeout=5000)

   buttons = page.locator("ul.jobs-search-pagination__pages li.jobs-search-pagination__indicator button")
   tot_buttons = buttons.count()
   print("Total Number of pages extracted: ",tot_buttons)
   url_list = []
   if tot_buttons != number_of_pages : #and number_of_pages<=1
      print("No Match")
   else: print("Math Matches")

   if tot_buttons > 1:
      for i in range(tot_buttons):
         try:
            print(f"Opening Paginated button in new window {i + 1}")

            # Click the button
            buttons.nth(i).click()

            # Wait for content to load (adjust selector for your use case)
            page.wait_for_timeout(5000)
            #page.wait_for_selector("div.jobs-search-results-list", timeout=5000)
            # Print current page information
            current_page = page.locator("div.jobs-search-pagination").inner_text()
            print(f"Currently on: {page.url}")

            # Extract jobs on the new page
            extract_job_ids(page)

         except Exception as e:
            print(f"Error clicking button {i + 1}: {e}")

   elif number_of_pages == 1:
      extract_job_ids(page)


   write_to_excel()
   applyJobs(page)
   print("Printing dataframe after autoapplying for jobs")
   #reading dataframe from excel sheet
   df = pd.read_excel(filename)
   print(df) #printing it.

def job_search(page, role, location):
   navbar = page.get_by_label("Global Navigation")
   navbar.is_visible()

   job_icon = page.get_by_role("link", name="Jobs", exact=True)
   job_icon.is_visible()
   job_icon.click()
   print('Clicked on Jobs icon')
   page.wait_for_timeout(2000)

   role_search = page.get_by_role("combobox", name="Search by title, skill, or")
   role_search.fill(role)
   print('Role filled')
   page.wait_for_timeout(2000)

   loc_search = page.get_by_role("combobox", name="City, state, or zip code")
   loc_search.fill(location)
   click_loc_search = page.get_by_label("Global Navigation").get_by_text(location, exact=True)
   click_loc_search.click()
   print('Location filled')
   page.wait_for_timeout(2000)

   job_posting_date = page.get_by_label("Date posted filter. Clicking")
   job_posting_date.click()
   page.wait_for_timeout(5000)
   date_option = page.get_by_text("Past 24 hours", exact=True)
   date_option.click()
   page.wait_for_timeout(5000)
   show_jobs = page.get_by_role("button", name="Apply current filter to show")
   show_jobs.click()
   print('Date Filter Applied')
   page.wait_for_timeout(5000)

   # Check if "No matching jobs found." message appears
   check = page.get_by_text("No matching jobs found.")
   if check.count() > 0:
      print("No Jobs found for the given constraints of role, location, and time of posting")
      return

   easy_apply = page.get_by_label("Easy Apply filter.")
   easy_apply.click()
   print('Easy Apply Clicked')
   page.wait_for_timeout(5000)

   # Check if "No matching jobs found." message appears
   check = page.get_by_text("No matching jobs found.")
   if check.count() > 0:
      print("No Jobs found for the given constraints of role, location, and time of posting")
      return

   extract_jobs(page)

def open_browser(playwright: Playwright, url: str):
   browser = playwright.chromium.launch(headless=head)
   page = browser.new_page(viewport={'width': 1600, 'height': 900})

   try:
      page.goto(url)
      page.wait_for_timeout(1000)

      email_input = page.get_by_label('Email or phone')
      email_input.fill(EMAIL)

      password_input = page.get_by_label('Password')
      password_input.fill(PASSWORD)

      checkbox = page.get_by_text("Keep me logged in")
      if checkbox.is_visible():
          checkbox.click()

      login_button = page.locator('button.btn__primary--large[type="submit"]')
      login_button.click()

      page.wait_for_timeout(5000)
      navbar = page.get_by_label("Global Navigation")
      if navbar.is_visible():
          print("Login successful!!!")
      else:
          print("Login unsuccessful!!")

      job_search(page, role, location)

   except Exception as e:
      print(f'Error occurred: {e}')
      print("Couldn't login to the application!")

   finally:
      browser.close()

def main():
   with sync_playwright() as playwright:
      open_browser(
         playwright=playwright,
         url=url
      )

if __name__ == '__main__':
   print("Running Automation Script")
   main()
   print("\nFinal Status Dictionary:", status_dict)
   update_excel_from_dict(filename)


